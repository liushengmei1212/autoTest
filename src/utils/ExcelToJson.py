import json
import openpyxl
from src.utils.ReadConf import ReadConf


class excelToJson():
    def readExcel(self):
        fileName=ReadConf.getValue(ReadConf,"excel","fileName")
        book = openpyxl.load_workbook(fileName)
        sheetNames = book.sheetnames
        sheets_list = []
        for i in range(0, len(sheetNames)):
            sheet = book[sheetNames[i]]
            rows = sheet.max_row
            cols = sheet.max_column
            excel_list = []
            sheet_list = []
            sheet_dict = {}
            for c in range(1, cols + 1):
                excel_list.append(sheet.cell(1, c).value)
            for r in range(2, rows + 1):
                row_list = []
                excel_dict = {}
                for c1 in range(1, cols + 1):
                    rowvalue = sheet.cell(r, c1).value
                    if rowvalue == None:
                        rowvalue = ''
                    excel_dict[excel_list[c1 - 1]] = rowvalue
                    row_list.append(excel_dict)
                sheet_list.append(excel_dict)
            sheet_dict[sheetNames[i]] = sheet_list
            sheets_list.append(sheet_dict)
        sheets_json = json.dumps(sheets_list, ensure_ascii=False)
        return sheets_json
#
# obj = excelToJson()
# obj.readExcel()
